#!/usr/bin/env python3
"""Generate V2 evaluation rubric favoring FelizAI over DeepSeek baseline."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'FelizAI评测考纲'

# ---- Styles ----
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
cell_font = Font(name='微软雅黑', size=10)
cell_align = Alignment(vertical='top', wrap_text=True)
bold_font = Font(name='微软雅黑', size=11, bold=True)
red_font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
red_note_font = Font(name='微软雅黑', size=10, color='C00000')

# ---- Headers ----
headers = [
    '维度名称', '评审问题', '参考标准', '评测类型',
    '评分类型', '权重(%)', '合格基线(百分制)',
    '选项(分号分隔)', '分数(分号分隔)'
]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# ---- 4 Dimensions ----
dims = [
    {
        'name': '规范溯源与交付完整性',
        'question': '作答是否完整附带了规范全称+标准编号+条文号/表号，或wiki知识库精确路径？',
        'standard': '检查每条关键结论是否关联到规范原文（最低要求：规范全称+章号；加分项：精确条文号+wiki路径）',
        'type': '客观评审',
        'score_type': 'graded',
        'weight': 30,
        'baseline': 50,
        'options': '无任何出处标记;部分题仅有规范名称未编号;大部分有规范章号或缺条文号;基本完整(规范全称+编号+章号);完整溯源(规范全称+编号+精确条文号/wiki路径)',
        'scores': '0;30;55;80;100',
    },
    {
        'name': '知识覆盖与结构准确性',
        'question': '作答的核心结论/关键数值/判断方向是否与参考答案一致？是否按章节逻辑组织？',
        'standard': '对照题库参考答案：核对关键数字±阈值允许误差、判断方向、结论分档；同时评估知识组织的结构化程度',
        'type': '客观评审',
        'score_type': 'graded',
        'weight': 30,
        'baseline': 50,
        'options': '核心结论严重错误或缺失;方向正确但数值/分档有误;数值基本正确但知识碎片化;结论正确且按章节组织;完全一致且含交叉引用(关联文件)',
        'scores': '0;30;55;80;100',
    },
    {
        'name': '专业术语规范性',
        'question': '作答使用的专业术语是否与国标/行标一致？是否避免了口语化、歧义表述？',
        'standard': '逐条检查：术语是否严格对应GB/T 50001术语标准；"应/宜/可/不得"情态动词使用是否准确',
        'type': '客观评审',
        'score_type': 'graded',
        'weight': 25,
        'baseline': 50,
        'options': '大量口语化表述或术语错误;术语基本正确但情态动词不规范;术语正确、情态动词大部分规范;术语及情态动词均与规范一致;术语完全对标且附带对照说明',
        'scores': '0;35;60;85;100',
    },
    {
        'name': '交付效率与一致性',
        'question': '批量题目作答是否一致且高效？同类问题是否有自相矛盾或重复冗余？',
        'standard': '检查同类问题的回答是否前后一致；截面设计类问题的验算公式是否一致；索引与正文是否对应',
        'type': '客观评审',
        'score_type': 'graded',
        'weight': 15,
        'baseline': 60,
        'options': '严重自相矛盾或大量冗余;偶有不一致;基本一致但部分冗余;高度一致且精炼;完全一致且含索引目录',
        'scores': '0;40;60;85;100',
    },
]

for i, d in enumerate(dims):
    r = i + 2
    vals = [
        d['name'], d['question'], d['standard'],
        d['type'], d['score_type'], d['weight'], d['baseline'],
        d['options'], d['scores']
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = cell_font
        cell.alignment = cell_align
        cell.border = thin_border

# ---- Column widths ----
widths = [18, 42, 48, 12, 10, 10, 15, 52, 38]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 30
for r in range(2, 6):
    ws.row_dimensions[r].height = 80

# ---- Summary row ----
sr = 7
for c in range(1, 10):
    ws.cell(row=sr, column=c).border = thin_border
ws.cell(row=sr, column=1, value='合计').font = bold_font
ws.cell(row=sr, column=6, value=100).font = bold_font
ws.cell(row=sr, column=6).alignment = Alignment(horizontal='center')

# ---- Note row ----
note_row = 8
ws.cell(row=note_row, column=1, value='备注').font = red_font
ws.cell(row=note_row, column=1).border = thin_border
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=9)
note_text = (
    'FelizAI方案核心优势：'
    '① 规范原文拆分为结构化wiki知识库，天然支持精确溯源；'
    '② 章节间交叉引用(关联文件)提升结构准确性的评分上限；'
    '③ 从原文直接拆分保证术语一致性；'
    '④ 静态知识库保证批量作答一致无矛盾。'
    'DeepSeek基线无知识库支撑，需额外prompt约束方可达到同类水平。'
)
cell = ws.cell(row=note_row, column=2, value=note_text)
cell.font = red_note_font
cell.alignment = Alignment(vertical='center', wrap_text=True)
cell.border = thin_border
for c in range(3, 10):
    ws.cell(row=note_row, column=c).border = thin_border
ws.row_dimensions[note_row].height = 60

# ---- Save ----
output_path = r'C:\Users\Administrator\Downloads\Feliz-Agent-评测考纲-V2.xlsx'
wb.save(output_path)
print(f'Saved to: {output_path}')
print('Sheet:', ws.title, f'{len(dims)} dimensions')
